import openpyxl
import datetime


def file_loc():
    file = filedialog.askopenfile(mode='r', filetypes=[('excel', '*.xlsx')])

    if file:
        file_path = os.path.abspath(file.name)
        file_name = os.path.basename(file_path)

        now = datetime.datetime.now()
        output_file_name = now.strftime("%B%d%X").replace(":", "")

        # Load the Excel workbook
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        max_row = sheet.max_row

        # rearrange the columns
        for i in range(1, max_row + 1):
            value = sheet.cell(row=i, column=5).value  # Column A is represented by index 1
            if value and isinstance(value, str):  # Check if the value is a non-empty string
                value = value.replace(" ", "")  # Remove spaces
                sheet.cell(row=i, column=5).value = value

            secret = sheet.cell(row=i, column=5).value  # Column E is represented by index 5
            anon = sheet.cell(row=i, column=7).value
            file = sheet.cell(row=i, column=9).value
            comp = sheet.cell(row=i, column=6).value

            sheet.cell(row=i, column=1).value = secret  # Column A is represented by index 1
            sheet.cell(row=i, column=2).value = anon
            sheet.cell(row=i, column=3).value = file
            sheet.cell(row=i, column=5).value = comp

            sheet.cell(row=i, column=6).value = None
            sheet.cell(row=i, column=7).value = None
            sheet.cell(row=i, column=8).value = None
            sheet.cell(row=i, column=9).value = None
            sheet.cell(row=i, column=10).value = None

        # clean the sheet
        for j in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
            cell = j[0]  # Get the cell in column A
            if cell.value and str(cell.value).startswith(" "):
                cell.value = str(cell.value)[1:]  
            if cell.value and str(cell.value).startswith("##"):
                cell.value = str(cell.value)[1:] 
            if cell.value and str(cell.value).startswith("##"):
                cell.value = str(cell.value)[2:]  
            if cell.value and str(cell.value).startswith("##"):
                cell.value = str(cell.value)[1:]  

        # Save the modified workbook
        workbook.save(f'{output_file_name}.xlsx')
        workbook.close()

        # User Download Folder
        dl_folder = os.path.expanduser('~') + '/Downloads/'
        open = f'{dl_folder}{output_file_name}.xlsx'
        new_path = os.path.join(dl_folder, f'{output_file_name}.xlsx')
        os.rename(f'{output_file_name}.xlsx', new_path)
        print(open)
        subprocess.Popen(['start', open])
